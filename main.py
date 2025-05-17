from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import random

wb = load_workbook('sticker-scheme.xlsx')
ws = wb.active

print("The Generated names are: ")
for i in range(5):
    rand_num = random.randint(2,ws.max_row)
    name = ws.cell(row=rand_num, column=1).value
    selected = ws.cell(row=rand_num, column=2,)
    selected.value = '✔️'
    print(f"{name}")

wb.save('sticker-scheme.xlsx')
